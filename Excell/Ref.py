from openpyxl import Workbook,load_workbook,styles

ref_wb = load_workbook("Referans.xlsx")
ref_ws = ref_wb.active

# Aktif çalışma sayfasının adını yazdırma
# print(wb.sheetnames)      # <Worksheet "İsimler">

#Sütun sayısını elinizle belirtmek yerine ws.max_row ve ws.max_column değişkenlerini kullanabilirsiniz.
# print(ref_key_list) 

for satir in range(1,ref_ws.max_row+1):
    ref_ws.cell(satir,2).value = str(ref_ws.cell(satir,1).value).partition(' ')[0]
    print(str(ref_ws.cell(satir,2).value))

ref_wb.save("Referans_2.xlsx")

# print(target_wb.sheetnames)     # ['İlk Çalışma Alanı', 'Posta Kodları', 'Ülkeler']