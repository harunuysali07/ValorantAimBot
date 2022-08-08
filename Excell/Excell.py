from openpyxl import Workbook,load_workbook,styles

HücreFont = styles.fonts.Font(bold=True)

wb = load_workbook("Source.xlsx")
ws = wb.active

ref_wb = load_workbook("Referans.xlsx")
ref_ws = ref_wb.active

ref_key_list = []

for satir in range(1,ref_ws.max_row+1):
    ref_key_list.append(str(ref_ws.cell(satir,2).value))

# Aktif çalışma sayfasının adını yazdırma
# print(wb.sheetnames)      # <Worksheet "İsimler">

#Sütun sayısını elinizle belirtmek yerine ws.max_row ve ws.max_column değişkenlerini kullanabilirsiniz.
# print(ref_key_list) 

for key in range(0, len(ref_key_list)):
    if not ref_key_list[key] == "None":
        for satir in range(1674,ws.max_row+1):
            if not str(ws.cell(satir,5).value) == "None":
                if not str(ws.cell(satir,5).value) == str(ref_ws.cell(key + 1,1).value):
                    if " " + ref_key_list[key] + " " in " " + str(ws.cell(satir,5).value) + " ":
                        if  ref_ws.cell(key + 1,3).value == ws.cell(satir,4).value:
                            print(str(ws.cell(satir,5).value) + " | " + str(ws.cell(satir,4).value) + " | " + "    --->>>   " + str(ref_ws.cell(key + 1,1).value))
                            ws.cell(satir,5).value = ref_ws.cell(key + 1,1).value
                            ws.cell(satir,5).fill = styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type = "solid")
                        elif  " " + str(ref_ws.cell(key + 1,3).value) + " " in " " + str(ws.cell(satir,4).value) + " ":
                            print(str(ws.cell(satir,5).value) + " | " + str(ws.cell(satir,4).value) + " | " + "    --->>>   " + str(ref_ws.cell(key + 1,1).value))
                            ws.cell(satir,5).value = ref_ws.cell(key + 1,1).value
                            ws.cell(satir,4).value = ref_ws.cell(key + 1,3).value
                            ws.cell(satir,5).fill = styles.PatternFill(start_color="00BBFF", end_color="00BBFF", fill_type = "solid")
                            ws.cell(satir,4).fill = styles.PatternFill(start_color="00BBFF", end_color="00BBFF", fill_type = "solid")
                else:
                    ws.cell(satir,5).fill = styles.PatternFill(start_color="FF00DD", end_color="FF00DD", fill_type = "solid")

wb.save("Result_7.xlsx")

# print(target_wb.sheetnames)     # ['İlk Çalışma Alanı', 'Posta Kodları', 'Ülkeler']